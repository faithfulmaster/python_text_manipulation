import openpyxl
import re
import os
import sys
import glob
import errno
import datetime

files = glob.glob('Source/1PE.xlsx')
for name in files:
    file_name = name.split('/')[1]
    print file_name

    # Open xlsx file
    wb1 = openpyxl.load_workbook(name)

    files = glob.glob('Urdu/601PETER.xlsx')
    for name in files:
        file_name = name.split('/')[1]
        print file_name

        # Open xlsx file
        wb2 = openpyxl.load_workbook(name)

        print "Conversion in progress"

        # Access each element of a column row by row and write it to md file
        sheet1 = wb1.get_sheet_by_name("Sheet1")
        for row1 in range(2, sheet1.max_row+1):

            text = sheet1['D' + str(row1)].value
            if text != None:
                text = text.encode('utf-8')
                text = text.split('\n')

                sheet2 = wb2.get_sheet_by_name("Sheet1")
                target = ""

                for i in text:
                    for row2 in range(3, sheet2.max_row+1):
                        text2 = sheet2['B' + str(row2)].value
                        transl = sheet2['D' + str(row2)].value

                        if type(text2) == unicode and transl != None:
                            text2 = text2.encode('utf-8')


                            if re.search(text2, i):
                                if transl[-1] == u"\u0964":
                                    transl = transl[0:-1]
                                    transl = transl.encode('utf-8')
                                    t = re.sub(text2, transl, i)
                                elif transl.endswith('?'):
                                    transl = transl.encode('utf-8')
                                    text2 = text2[0:-1]
                                    text2 = text2 + '\?'
                                    transl = transl[0:-1]
                                    transl = transl + '?'
                                    t = re.sub(text2, transl, i)
                                else:
                                    transl = transl.encode('utf-8')
                                    t = re.sub(text2, transl, i)
                                t = t.decode('utf-8') + "\n"
                                target = target + t

                sheet1['E' + str(row1)] = target.strip()

    # Close the file
    wb2.close()
    wb1.save(filename=file_name)
    wb1.close()
    print 'Conversion Done !'
