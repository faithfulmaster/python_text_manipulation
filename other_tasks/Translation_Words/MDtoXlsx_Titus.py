#!/usr/bin/python
# -*- coding: utf-8 -*-

import openpyxl
import re
import os
import sys
import glob
import errno

# Open xlsx file
files = glob.glob('**/*.xlsx')
for name in files:
    file_name = name.split('/')[1]
    print file_name

    wb1 = openpyxl.load_workbook(name)
    wb2 = openpyxl.Workbook()
    
    print "Conversion in progress !"

    # Access each element of a column row by row and write it to md file
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    sheet2 = wb2.get_sheet_by_name("Sheet")

    mdfiles = glob.glob('**/**/*.md')
    prev_word = ""
    row2 = 1

    sheet2['A' + str(row2)] = "Translation Word"
    sheet2['B' + str(row2)] = "Words"
    sheet2['C' + str(row2)] = "Definition"
    sheet2['D' + str(row2)] = "Definition_hi"
    sheet2['E' + str(row2)] = "Facts"
    sheet2['F' + str(row2)] = "Facts_hi"
    sheet2['G' + str(row2)] = "Bible References"
    sheet2['H' + str(row2)] = "Bible References_hi"
    sheet2['I' + str(row2)] = "Translation Suggestions"
    sheet2['J' + str(row2)] = "Translation Suggestions_hi"
    sheet2['K' + str(row2)] = "Bible Examples"
    sheet2['L' + str(row2)] = "Bible Examples_hi"

    for row1 in range(1, sheet1.max_row+1):

		for i in mdfiles:
			filename = i.split('/')[2]
			fn = filename.split('.')[0]

			word = sheet1['A' + str(row1)].value
			if word != None:
				if prev_word != word:
					if word == fn:
						row2 += 1
						sheet2['A' + str(row2)] = word
						column = ""
						addline = ""
						prev_word = word
						with open(i) as inputfile:
							for inpline in inputfile:
								process1 = re.sub("\(\.\.(\/*\w*\d*)*\.md\)", "", inpline)
								process2 = re.sub("\(rc:(\/*\w*-*\d*)*\)", "", process1)
								line = re.sub("[HG]\d+\s?,?", "", process2)
								if re.search(r"##\s?Definition:\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "C"
								elif re.search(r"##\s?Facts:\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "E"
								elif re.search(r"##\s?Bible References:\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "G"
								elif re.search(r"##\s?Translation Suggestions:?\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "I"
								elif re.search(r"##\s?Examples from the Bible stories:\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "K"
								elif re.search(r"##\s?Word Data:?\s?", line):
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = "M"
								elif re.search(r"#(\s*\w*,*'*-*\(*\)*)*", line) and line[1] != "#":
									if column != "":
										sheet2[column + str(row2)] = addline.strip()
										addline = ""
									column = ""
									addline = ""
									sheet2['B' + str(row2)] = str(line[2:-3])

								else:
									addline += line
									if column == "":
										print line
							sheet2[column + str(row2)] = addline.strip()

    # Close the file
    fil = "New.xlsx"
    wb2.save(filename=fil)
    wb2.close()
    wb1.close()
    print 'Conversion Done !'
