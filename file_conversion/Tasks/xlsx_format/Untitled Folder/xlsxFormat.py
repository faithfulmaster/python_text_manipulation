#!/usr/bin/python
# -*- coding: utf-8 -*-

import openpyxl
import re
import os
import sys
import glob
import errno

files = glob.glob('**/*.xlsx')
for name in files:
    file_name = name.split('/')[1]
    print file_name
    print "Conversion in progress"

    # Open xlsx file
    wb1 = openpyxl.load_workbook(name)
    wb2 = openpyxl.Workbook()

    # Access each element of a column row by row and write it to md file
    sheet1 = wb1.get_sheet_by_name("Sheet1")
    sheet2 = wb2.get_sheet_by_name("Sheet")
    row2 = 2;
    for row1 in range(2, sheet1.max_row+1):

        try:
            # Read bookname
            bookname = sheet1['A' + str(row1)].value

            # Read chapter
            chapter = sheet1['B' + str(row1)].value
            if chapter != None:
                ch = chapter.split('-')
                chapter = ch[0]

            # Read verse and text
            text = sheet1['C' + str(row1)].value

            result = re.search(r"verse number=\"(\d+(-\d+)?)\" style=\"v\" />(.+?)<verse", text)
            while result != None:
                verse = result.group(1)
                content = result.group(3)
                sheet2['A' + str(row2)] = bookname
                sheet2['B' + str(row2)] = chapter
                sheet2['C' + str(row2)] = verse
                sheet2['D' + str(row2)] = content.strip()
                row2 += 1
                text = text[:result.start()] + text[result.end()-5:]
                result = re.search(r"verse number=\"(\d+(-\d+)?)\" style=\"v\" />(.+?)<verse", text)

            if result == None:
                result = re.search(r"verse number=\"(\d+(-\d+)?)\" style=\"v\" />(.+)", text)
                verse = result.group(1)
                content = result.group(3)
                sheet2['A' + str(row2)] = bookname
                sheet2['B' + str(row2)] = chapter
                sheet2['C' + str(row2)] = verse
                sheet2['D' + str(row2)] = content.strip()
                row2 += 1

        except:
            print "Error found at row " + str(row1)

    # Close the file
    wb1.close()
    wb2.save(filename=file_name)
    wb2.close()
    print 'Conversion Done !'
